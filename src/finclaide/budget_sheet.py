from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from finclaide.category_filters import is_ynab_system_category
from finclaide.database import Database, utc_now
from finclaide.plan_service import (
    insert_plan_revision,
    read_plan_categories_snapshot,
)
from finclaide.errors import DataIntegrityError
from finclaide.money import to_milliunits
from finclaide.months import parse_due_month


INFLOW_GROUP_NAMES = frozenset({"Monthly Income", "Yearly Income"})

# Match a contiguous SUM range like =SUM(F10:F61) so the validator can
# diagnose which parsed rows fall outside the cached SUM and surface them in
# the error.
_SUM_RANGE_PATTERN = re.compile(
    r"=\s*SUM\s*\(\s*([A-Z]+)(\d+)\s*:\s*([A-Z]+)(\d+)\s*\)",
    re.IGNORECASE,
)
_SOURCE_CELL_PATTERN = re.compile(r"^([A-Z]+)(\d+)$")


@dataclass(frozen=True)
class _TotalCellInfo:
    cached_value: Any
    cell_ref: str
    formula_text: str | None


def _kind_for_group(group_name: str) -> str:
    return "inflow" if group_name in INFLOW_GROUP_NAMES else "outflow"


@dataclass(frozen=True)
class PlannedCategoryRow:
    group_name: str
    category_name: str
    block: str
    source_cell: str
    planned_milliunits: int
    annual_target_milliunits: int
    due_month: int | None
    formula_text: str | None
    kind: str = "outflow"


class BudgetImporter:
    TOTAL_TOLERANCE_MILLIUNITS = 10

    def __init__(self, database: Database) -> None:
        self.database = database

    def import_budget(self, workbook_path: Path, sheet_name: str) -> dict[str, object]:
        workbook_path = Path(workbook_path)
        if not workbook_path.exists():
            raise DataIntegrityError(f"Workbook not found at {workbook_path}")

        workbook_formula = load_workbook(workbook_path, data_only=False)
        workbook_cached = load_workbook(workbook_path, data_only=True)
        if sheet_name not in workbook_formula.sheetnames:
            raise DataIntegrityError(f"Required sheet '{sheet_name}' not found in workbook.")

        sheet_formula = workbook_formula[sheet_name]
        sheet_cached = workbook_cached[sheet_name]
        max_row = sheet_formula.max_row

        plan_year = self._extract_plan_year(sheet_name)
        rows = []
        rows.extend(self._parse_monthly_block(sheet_formula, sheet_cached))
        rows.extend(self._parse_yearly_block(sheet_formula, sheet_cached))
        rows.extend(self._parse_single_group_block(sheet_formula, sheet_cached, "I", "J", 2, max_row, "Stipends"))
        rows.extend(self._parse_single_group_block(sheet_formula, sheet_cached, "L", "M", 2, max_row, "Savings"))

        duplicates = self._find_duplicates(rows)
        if duplicates:
            raise DataIntegrityError(
                f"Duplicate planned categories detected: {', '.join(sorted(duplicates))}"
            )

        self._validate_totals(sheet_formula, sheet_cached, rows)

        digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
        summary = {
            "imported_at": utc_now(),
            "plan_year": plan_year,
            "row_count": len(rows),
            "groups": sorted({row.group_name for row in rows}),
        }

        with self.database.connect() as connection:
            cursor = connection.execute(
                """
                INSERT INTO budget_imports(workbook_path, workbook_sha256, sheet_name, imported_at, plan_year, summary_json)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    str(workbook_path),
                    digest,
                    sheet_name,
                    summary["imported_at"],
                    plan_year,
                    json.dumps(summary, sort_keys=True),
                ),
            )
            import_id = int(cursor.lastrowid)
            group_ids: dict[str, int] = {}
            for group_name, block in sorted({(row.group_name, row.block) for row in rows}):
                group_cursor = connection.execute(
                    """
                    INSERT INTO planned_groups(import_id, name, block)
                    VALUES (?, ?, ?)
                    """,
                    (import_id, group_name, block),
                )
                group_ids[group_name] = int(group_cursor.lastrowid)

            for row in rows:
                connection.execute(
                    """
                    INSERT INTO planned_categories(
                        import_id,
                        planned_group_id,
                        group_name,
                        category_name,
                        block,
                        source_cell,
                        planned_milliunits,
                        annual_target_milliunits,
                        due_month,
                        formula_text
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        import_id,
                        group_ids[row.group_name],
                        row.group_name,
                        row.category_name,
                        row.block,
                        row.source_cell,
                        row.planned_milliunits,
                        row.annual_target_milliunits,
                        row.due_month,
                        row.formula_text,
                    ),
                )

            new_plan_id = self._mirror_into_plan_model(
                connection,
                plan_year=plan_year,
                plan_name=sheet_name,
                rows=rows,
                source_import_id=import_id,
            )

        summary["import_id"] = import_id
        summary["plan_id"] = new_plan_id
        return summary

    def _mirror_into_plan_model(
        self,
        connection,
        *,
        plan_year: int,
        plan_name: str,
        rows: list[PlannedCategoryRow],
        source_import_id: int,
    ) -> int:
        now = utc_now()
        # Capture the about-to-be-archived plan's categories so an operator
        # who had in-flight edits can restore them after the import. Tagged
        # with the new plan_id below so the restore replaces the new plan's
        # categories (the active plan) — closes the lost-edit window.
        old_active_row = connection.execute(
            "SELECT id FROM plans WHERE plan_year = ? AND status = 'active'",
            (plan_year,),
        ).fetchone()
        pre_archive_snapshot: list[dict[str, Any]] = []
        if old_active_row is not None:
            pre_archive_snapshot = read_plan_categories_snapshot(
                connection, int(old_active_row["id"])
            )
        connection.execute(
            """
            UPDATE plans
            SET status = 'archived', archived_at = ?, updated_at = ?
            WHERE plan_year = ? AND status = 'active'
            """,
            (now, now, plan_year),
        )
        plan_cursor = connection.execute(
            """
            INSERT INTO plans(
                plan_year, name, status, source,
                created_at, updated_at, source_import_id
            )
            VALUES (?, ?, 'active', 'imported', ?, ?, ?)
            """,
            (plan_year, plan_name, now, now, source_import_id),
        )
        new_plan_id = int(plan_cursor.lastrowid)
        if pre_archive_snapshot:
            insert_plan_revision(
                connection,
                plan_id=new_plan_id,
                source="importer",
                summary=(
                    f"Importer overwrote plan ({len(pre_archive_snapshot)} "
                    f"categories saved for restore)"
                ),
                change_count=len(pre_archive_snapshot),
                snapshot=pre_archive_snapshot,
                when=now,
            )
        for row in rows:
            connection.execute(
                """
                INSERT INTO plan_categories(
                    plan_id, group_name, category_name, block, kind,
                    planned_milliunits, annual_target_milliunits, due_month,
                    notes, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?)
                """,
                (
                    new_plan_id,
                    row.group_name,
                    row.category_name,
                    row.block,
                    row.kind,
                    row.planned_milliunits,
                    row.annual_target_milliunits,
                    row.due_month,
                    now,
                    now,
                ),
            )
        return new_plan_id

    def _parse_monthly_block(self, sheet_formula, sheet_cached) -> list[PlannedCategoryRow]:
        rows: list[PlannedCategoryRow] = []
        current_group: str | None = None
        first_data_row = self._formula_range_start_row(sheet_formula, "A", "B", "B53") or 2
        # Iterate from row 1 so a 'Monthly Income' header in A1 can register as
        # a group. Outflow rows must still be inside the SUM range so the
        # cached total reconciles in _validate_totals; inflow rows under a
        # known income group can live above the SUM range (which is how
        # current sheets store paychecks).
        for row_number in range(1, sheet_formula.max_row + 1):
            name_cell = f"A{row_number}"
            amount_cell = f"B{row_number}"
            name_value = self._text_value(sheet_formula[name_cell].value)
            amount_value = sheet_cached[amount_cell].value
            if not name_value:
                continue
            if self._is_summary_label(name_value):
                break
            if amount_value in (None, ""):
                current_group = name_value
                continue
            is_inflow_group = current_group in INFLOW_GROUP_NAMES
            if row_number < first_data_row and not is_inflow_group:
                continue
            if current_group is None:
                continue
            if is_ynab_system_category(current_group, name_value):
                continue
            formula_text = self._formula_text(sheet_formula[amount_cell].value)
            rows.append(
                PlannedCategoryRow(
                    group_name=current_group,
                    category_name=name_value,
                    block="monthly",
                    source_cell=amount_cell,
                    planned_milliunits=to_milliunits(amount_value),
                    annual_target_milliunits=to_milliunits(amount_value),
                    due_month=None,
                    formula_text=formula_text,
                    kind=_kind_for_group(current_group),
                )
            )
        return rows

    def _parse_yearly_block(self, sheet_formula, sheet_cached) -> list[PlannedCategoryRow]:
        rows: list[PlannedCategoryRow] = []
        current_group: str | None = None
        # Iterate from row 1 so a 'Yearly Income' header in D1 can register
        # as a group. Without this, irregular inflows (Bonus, Tax Return,
        # etc.) get filtered out and never reach the cascade.
        for row_number in range(1, sheet_formula.max_row + 1):
            name_cell = f"D{row_number}"
            total_cell = f"E{row_number}"
            due_cell = f"F{row_number}"
            monthly_cell = f"G{row_number}"
            name_value = self._text_value(sheet_formula[name_cell].value)
            total_value = sheet_cached[total_cell].value
            due_value = self._text_value(sheet_formula[due_cell].value)
            due_or_monthly_value = sheet_cached[due_cell].value
            monthly_value = sheet_cached[monthly_cell].value
            if not name_value:
                continue
            if self._is_summary_label(name_value):
                break
            if self._is_yearly_group_header(total_value, due_or_monthly_value, monthly_value):
                current_group = name_value
                continue
            if current_group is None:
                continue
            planned_value, source_cell = self._resolve_yearly_planned_value(
                row_number=row_number,
                due_or_monthly_value=due_or_monthly_value,
                monthly_value=monthly_value,
            )
            if is_ynab_system_category(current_group, name_value):
                continue
            rows.append(
                PlannedCategoryRow(
                    group_name=current_group,
                    category_name=name_value,
                    block="annual" if current_group == "Yearly" else "one_time",
                    source_cell=source_cell,
                    planned_milliunits=to_milliunits(planned_value),
                    annual_target_milliunits=to_milliunits(total_value),
                    due_month=parse_due_month(due_value) or parse_due_month(name_value),
                    formula_text=self._formula_text(sheet_formula[source_cell].value),
                    kind=_kind_for_group(current_group),
                )
            )
        return rows

    def _parse_single_group_block(
        self,
        sheet_formula,
        sheet_cached,
        name_column: str,
        amount_column: str,
        start_row: int,
        end_row: int,
        expected_group: str,
    ) -> list[PlannedCategoryRow]:
        name_column = self._find_header_column(sheet_formula, expected_group, [name_column, chr(ord(name_column) - 1)])
        amount_column = get_column_letter(ord(name_column) - 64 + 1)
        group_value = self._text_value(sheet_formula[f"{name_column}1"].value)
        rows: list[PlannedCategoryRow] = []
        current_group: str | None = group_value
        for row_number in range(start_row, end_row + 1):
            name_cell = f"{name_column}{row_number}"
            amount_cell = f"{amount_column}{row_number}"
            name_value = self._text_value(sheet_formula[name_cell].value)
            amount_value = sheet_cached[amount_cell].value
            if not name_value:
                continue
            if self._is_summary_label(name_value):
                break
            if amount_value in (None, ""):
                current_group = name_value
                continue
            if current_group is None:
                raise DataIntegrityError(f"Missing group header before {name_cell}.")
            if is_ynab_system_category(current_group, name_value):
                continue
            rows.append(
                PlannedCategoryRow(
                    group_name=current_group,
                    category_name=name_value,
                    block=expected_group.lower(),
                    source_cell=amount_cell,
                    planned_milliunits=to_milliunits(amount_value),
                    annual_target_milliunits=to_milliunits(amount_value),
                    due_month=None,
                    formula_text=self._formula_text(sheet_formula[amount_cell].value),
                    kind=_kind_for_group(current_group),
                )
            )
        return rows

    def _validate_totals(
        self,
        sheet_formula,
        sheet_cached,
        rows: list[PlannedCategoryRow],
    ) -> None:
        # Compare cached SUM cells against outflow rows only. Income rows
        # under 'Monthly Income' / 'Yearly Income' groups live above the
        # SUM range in the source sheet and so are not part of those totals.
        outflow_rows = [row for row in rows if row.kind == "outflow"]
        block_specs = {
            "monthly": {
                "row_blocks": {"monthly"},
                "label_columns": ["A"],
                "amount_columns": ["B"],
                "legacy_cell": "B53",
            },
            "annual": {
                "row_blocks": {"annual", "one_time"},
                "label_columns": ["D", "E", "F"],
                "amount_columns": ["E", "F", "G"],
                "legacy_cell": "G53",
            },
            "stipends": {
                "row_blocks": {"stipends"},
                "label_columns": ["H", "I"],
                "amount_columns": ["I", "J"],
                "legacy_cell": "J53",
            },
            "savings": {
                "row_blocks": {"savings"},
                "label_columns": ["K", "L"],
                "amount_columns": ["L", "M"],
                "legacy_cell": "M53",
            },
        }
        mismatches: list[str] = []
        for name, spec in block_specs.items():
            block_rows = [row for row in outflow_rows if row.block in spec["row_blocks"]]
            parsed_total = sum(row.planned_milliunits for row in block_rows)
            total_info = self._total_cell_info(
                sheet_formula,
                sheet_cached,
                label_columns=spec["label_columns"],
                amount_columns=spec["amount_columns"],
                legacy_cell=spec["legacy_cell"],
            )
            expected_total = to_milliunits(total_info.cached_value)
            if abs(parsed_total - expected_total) <= self.TOTAL_TOLERANCE_MILLIUNITS:
                continue
            mismatches.append(
                self._format_total_mismatch(
                    block_name=name,
                    parsed_total=parsed_total,
                    expected_total=expected_total,
                    total_info=total_info,
                    block_rows=block_rows,
                )
            )
        if mismatches:
            raise DataIntegrityError(
                "Budget totals do not match cached formulas:\n  " + "\n  ".join(mismatches)
            )

    def _format_total_mismatch(
        self,
        *,
        block_name: str,
        parsed_total: int,
        expected_total: int,
        total_info: "_TotalCellInfo",
        block_rows: list[PlannedCategoryRow],
    ) -> str:
        diff_milli = parsed_total - expected_total
        sum_range = self._extract_sum_range(total_info.formula_text)
        out_of_range = (
            self._rows_outside_range(block_rows, sum_range) if sum_range else []
        )
        cached_dollars = total_info.cached_value
        parsed_dollars = parsed_total / 1000
        diff_dollars = diff_milli / 1000
        formula_label = (
            f"{total_info.cell_ref} {total_info.formula_text or '(no formula)'}"
        )
        msg = (
            f"{block_name}: cached {formula_label} = ${cached_dollars:,.2f} "
            f"vs parsed ${parsed_dollars:,.2f} "
            f"(diff ${diff_dollars:+,.2f})."
        )
        if out_of_range:
            details = "; ".join(
                f"{row.source_cell} {row.group_name}/{row.category_name} "
                f"(${row.planned_milliunits / 1000:,.2f}/mo)"
                for row in out_of_range
            )
            msg += (
                f" Parsed rows outside the SUM range: {details}. "
                f"Fix: widen the SUM in {total_info.cell_ref} to include them, "
                f"or move them out of the block column."
            )
        elif sum_range is None:
            msg += " (Could not parse SUM range from the cached formula.)"
        return msg

    @staticmethod
    def _extract_sum_range(formula: str | None) -> tuple[str, int, str, int] | None:
        if not formula:
            return None
        match = _SUM_RANGE_PATTERN.search(formula)
        if not match:
            return None
        start_col, start_row, end_col, end_row = match.groups()
        try:
            return start_col, int(start_row), end_col, int(end_row)
        except ValueError:
            return None

    @staticmethod
    def _rows_outside_range(
        block_rows: list[PlannedCategoryRow],
        sum_range: tuple[str, int, str, int],
    ) -> list[PlannedCategoryRow]:
        _, start_row, _, end_row = sum_range
        offenders: list[PlannedCategoryRow] = []
        for row in block_rows:
            match = _SOURCE_CELL_PATTERN.match(row.source_cell or "")
            if not match:
                continue
            row_number = int(match.group(2))
            if row_number < start_row or row_number > end_row:
                offenders.append(row)
        return offenders

    def _total_cell_info(
        self,
        sheet_formula,
        sheet_cached,
        *,
        label_columns: list[str],
        amount_columns: list[str],
        legacy_cell: str,
    ) -> "_TotalCellInfo":
        for row_number in range(1, sheet_formula.max_row + 1):
            if not any(
                self._text_value(sheet_formula[f"{column}{row_number}"].value) == "Total"
                for column in label_columns
            ):
                continue
            for amount_column in amount_columns:
                cell_ref = f"{amount_column}{row_number}"
                formula_value = sheet_formula[cell_ref].value
                cached_value = sheet_cached[cell_ref].value
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    if cached_value is None:
                        raise DataIntegrityError(
                            f"Missing cached formula result in {cell_ref}."
                        )
                    return _TotalCellInfo(
                        cached_value=cached_value,
                        cell_ref=cell_ref,
                        formula_text=formula_value,
                    )
        # Fall back to the legacy fixed cell.
        cached = self._required_cached_value(sheet_formula, sheet_cached, legacy_cell)
        formula = sheet_formula[legacy_cell].value
        return _TotalCellInfo(
            cached_value=cached,
            cell_ref=legacy_cell,
            formula_text=formula if isinstance(formula, str) else None,
        )

    def _required_total_value(
        self,
        sheet_formula,
        sheet_cached,
        *,
        label_columns: list[str],
        amount_columns: list[str],
        legacy_cell: str,
    ):
        for row_number in range(1, sheet_formula.max_row + 1):
            if not any(
                self._text_value(sheet_formula[f"{column}{row_number}"].value) == "Total"
                for column in label_columns
            ):
                continue
            for amount_column in amount_columns:
                cell_ref = f"{amount_column}{row_number}"
                formula_value = sheet_formula[cell_ref].value
                cached_value = sheet_cached[cell_ref].value
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    if cached_value is None:
                        raise DataIntegrityError(f"Missing cached formula result in {cell_ref}.")
                    return cached_value
        return self._required_cached_value(sheet_formula, sheet_cached, legacy_cell)

    def _required_cached_value(self, sheet_formula, sheet_cached, cell_ref: str):
        formula_value = sheet_formula[cell_ref].value
        cached_value = sheet_cached[cell_ref].value
        if formula_value is None or not isinstance(formula_value, str) or not formula_value.startswith("="):
            raise DataIntegrityError(f"Expected formula in {cell_ref}.")
        if cached_value is None:
            raise DataIntegrityError(f"Missing cached formula result in {cell_ref}.")
        return cached_value

    def _formula_text(self, cell_value) -> str | None:
        if isinstance(cell_value, str) and cell_value.startswith("="):
            return cell_value
        return None

    def _formula_range_start_row(
        self,
        sheet_formula,
        label_column: str,
        amount_column: str,
        legacy_cell: str,
    ) -> int | None:
        for row_number in range(1, sheet_formula.max_row + 1):
            if self._text_value(sheet_formula[f"{label_column}{row_number}"].value) != "Total":
                continue
            formula_value = sheet_formula[f"{amount_column}{row_number}"].value
            start_row = self._extract_formula_range_start_row(formula_value, amount_column)
            if start_row is not None:
                return start_row
        return self._extract_formula_range_start_row(sheet_formula[legacy_cell].value, amount_column)

    def _extract_formula_range_start_row(self, formula_value: Any, amount_column: str) -> int | None:
        if not isinstance(formula_value, str) or not formula_value.startswith("="):
            return None
        match = re.search(rf"{amount_column}(\d+):{amount_column}\d+", formula_value)
        if match is None:
            return None
        return int(match.group(1))

    def _find_header_column(self, sheet_formula, expected_group: str, candidates: list[str]) -> str:
        for column in candidates:
            if self._text_value(sheet_formula[f"{column}1"].value) == expected_group:
                return column
        raise DataIntegrityError(f"Expected {expected_group!r} header in row 1.")

    def _is_yearly_group_header(self, total_value: Any, due_or_monthly_value: Any, monthly_value: Any) -> bool:
        return (
            self._numeric_value(total_value) is None
            and self._numeric_value(due_or_monthly_value) is None
            and self._numeric_value(monthly_value) is None
        )

    def _resolve_yearly_planned_value(
        self,
        *,
        row_number: int,
        due_or_monthly_value: Any,
        monthly_value: Any,
    ) -> tuple[Any, str]:
        if self._numeric_value(monthly_value) is not None:
            return monthly_value, f"G{row_number}"
        if self._numeric_value(due_or_monthly_value) is not None:
            return due_or_monthly_value, f"F{row_number}"
        raise DataIntegrityError(f"Missing annual planned amount at F{row_number}.")

    def _numeric_value(self, value: Any) -> Any:
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return value
        return None

    def _text_value(self, value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _is_summary_label(self, value: str) -> bool:
        return value in {"Total", "Remaining"}

    def _find_duplicates(self, rows: list[PlannedCategoryRow]) -> set[str]:
        seen: set[tuple[str, str]] = set()
        duplicates: set[str] = set()
        for row in rows:
            key = (row.group_name, row.category_name)
            if key in seen:
                duplicates.add(f"{row.group_name}/{row.category_name}")
            seen.add(key)
        return duplicates

    def _extract_plan_year(self, sheet_name: str) -> int:
        match = re.search(r"(20\d{2})", sheet_name)
        if not match:
            raise DataIntegrityError(f"Could not infer plan year from sheet name {sheet_name!r}.")
        return int(match.group(1))

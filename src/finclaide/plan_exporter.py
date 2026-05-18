from __future__ import annotations

import re
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from finclaide.category_filters import is_ynab_system_category
from finclaide.database import utc_now
from finclaide.plan_service import PlanService

# The importer in `budget_sheet.py` reads a fixed column layout. The
# exporter must match it cell-for-cell so the round-trip stays clean.
#
#   Monthly  → A:B    (group rows have empty B)
#   Yearly   → D:G    (group rows have empty E/F/G; G = monthly value;
#                      F = due month text; E = annual target)
#   Stipends → I:J    (group rows have empty J)
#   Savings  → L:M    (group rows have empty M)
#
# Totals live at row 53 (`B53`, `G53`, `J53`, `M53`) — matches the
# legacy fallback in `budget_sheet.py:369-397`. Totals carry `=SUM(...)`
# formulas with cached numeric values injected so the importer's
# `_required_total_value` validation passes.
TOTALS_ROW = 53

_MONTH_ABBREVIATIONS = (
    "",
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
)

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", _NS_MAIN)


@dataclass
class ExportResult:
    bytes: bytes
    filename: str
    row_count: int


@dataclass
class PlanCellGrid:
    """Layout-aware view of the active plan, decoupled from any output
    format. Both the xlsx exporter and the Sheets publisher render from
    this same grid so the column layout stays in lockstep."""

    cells: dict[str, str | float]
    """{cell_ref: value}. Cells with formula strings start with '='."""

    cached_values: dict[str, float]
    """{cell_ref: numeric_cached_value} for any cells that hold formulas
    requiring a cached numeric (the totals row's =SUM)."""

    row_count: int
    """Count of category rows actually written (excludes group headers
    and totals)."""


def build_plan_cell_grid(plan: dict[str, Any]) -> PlanCellGrid:
    cells: dict[str, str | float] = {}
    plan_blocks = {
        block: [
            row
            for row in plan["blocks"][block]
            if not is_ynab_system_category(
                row.get("group_name"),
                row.get("category_name"),
            )
        ]
        for block in ("monthly", "annual", "one_time", "stipends", "savings")
    }

    monthly_total = _emit_simple_block(
        cells,
        rows=plan_blocks["monthly"],
        name_col="A",
        amount_col="B",
        header_label=None,
    )
    cells[f"A{TOTALS_ROW}"] = "Total"
    cells[f"B{TOTALS_ROW}"] = "=SUM(B2:B52)"

    yearly_total = _emit_yearly_block(
        cells,
        annual_rows=plan_blocks["annual"],
        one_time_rows=plan_blocks["one_time"],
    )
    cells[f"D{TOTALS_ROW}"] = "Total"
    cells[f"G{TOTALS_ROW}"] = "=SUM(G2:G52)"

    stipends_total = _emit_simple_block(
        cells,
        rows=plan_blocks["stipends"],
        name_col="I",
        amount_col="J",
        header_label="Stipends",
    )
    cells[f"I{TOTALS_ROW}"] = "Total"
    cells[f"J{TOTALS_ROW}"] = "=SUM(J2:J52)"

    savings_total = _emit_simple_block(
        cells,
        rows=plan_blocks["savings"],
        name_col="L",
        amount_col="M",
        header_label="Savings",
    )
    cells[f"L{TOTALS_ROW}"] = "Total"
    cells[f"M{TOTALS_ROW}"] = "=SUM(M2:M52)"

    cached_values = {
        f"B{TOTALS_ROW}": _milliunits_to_dollars(monthly_total),
        f"G{TOTALS_ROW}": _milliunits_to_dollars(yearly_total),
        f"J{TOTALS_ROW}": _milliunits_to_dollars(stipends_total),
        f"M{TOTALS_ROW}": _milliunits_to_dollars(savings_total),
    }

    row_count = (
        len(plan_blocks["monthly"])
        + len(plan_blocks["annual"])
        + len(plan_blocks["one_time"])
        + len(plan_blocks["stipends"])
        + len(plan_blocks["savings"])
    )

    return PlanCellGrid(
        cells=cells, cached_values=cached_values, row_count=row_count
    )


def _emit_simple_block(
    cells: dict[str, str | float],
    *,
    rows: list[dict[str, Any]],
    name_col: str,
    amount_col: str,
    header_label: str | None,
) -> int:
    if header_label is not None:
        cells[f"{name_col}1"] = header_label
    next_row = 2
    current_group: str | None = None
    block_total = 0
    for row in rows:
        if row["group_name"] != current_group:
            cells[f"{name_col}{next_row}"] = row["group_name"]
            next_row += 1
            current_group = row["group_name"]
        cells[f"{name_col}{next_row}"] = row["category_name"]
        cells[f"{amount_col}{next_row}"] = _milliunits_to_dollars(row["planned_milliunits"])
        block_total += int(row["planned_milliunits"])
        next_row += 1
    return block_total


def _emit_yearly_block(
    cells: dict[str, str | float],
    *,
    annual_rows: list[dict[str, Any]],
    one_time_rows: list[dict[str, Any]],
) -> int:
    next_row = 2
    block_total = 0
    for current_group, group_rows in (
        ("Yearly", annual_rows),
        ("One Time Purchase", one_time_rows),
    ):
        if not group_rows:
            continue
        cells[f"D{next_row}"] = current_group
        next_row += 1
        for row in group_rows:
            cells[f"D{next_row}"] = row["category_name"]
            cells[f"E{next_row}"] = _milliunits_to_dollars(
                row.get("annual_target_milliunits", 0)
            )
            due_month = row.get("due_month")
            if due_month:
                cells[f"F{next_row}"] = _MONTH_ABBREVIATIONS[int(due_month)]
            cells[f"G{next_row}"] = _milliunits_to_dollars(row["planned_milliunits"])
            block_total += int(row["planned_milliunits"])
            next_row += 1
    return block_total


class PlanExporter:
    """Renders the active plan as an .xlsx matching `BudgetImporter`'s layout."""

    def __init__(self, plan_service: PlanService):
        self._plan_service = plan_service

    def export_active_plan(
        self,
        *,
        sheet_name: str,
        plan_year: int | None = None,
        now: str | None = None,
    ) -> ExportResult:
        plan = self._plan_service.get_active_plan(plan_year=plan_year)
        grid = build_plan_cell_grid(plan)
        timestamp = now or utc_now()

        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None  # openpyxl always creates one
        worksheet.title = sheet_name

        for cell_ref, value in grid.cells.items():
            worksheet[cell_ref] = value

        # Currency format on amount columns. The totals row's text label
        # cells (A53/D53/I53/L53) get the format too but it's harmless
        # since they hold strings.
        for column_letter in ("B", "E", "G", "J", "M"):
            for row in range(2, TOTALS_ROW + 1):
                cell = worksheet[f"{column_letter}{row}"]
                cell.number_format = "$#,##0.00"

        buffer = BytesIO()
        workbook.save(buffer)
        rendered_bytes = buffer.getvalue()

        # openpyxl can't compute formulas; inject cached values for the
        # totals so the importer's `_required_total_value` check passes
        # on a round-trip.
        rendered_bytes = _inject_cached_values(
            rendered_bytes, sheet_name, grid.cached_values
        )

        plan_name = plan["plan"]["name"]
        suggested_filename = _build_suggested_filename(plan_name, timestamp)

        return ExportResult(
            bytes=rendered_bytes,
            filename=suggested_filename,
            row_count=grid.row_count,
        )


def _milliunits_to_dollars(milliunits: int | None) -> float:
    if milliunits is None:
        return 0.0
    return round(int(milliunits) / 1000, 2)


def _build_suggested_filename(plan_name: str, timestamp: str) -> str:
    # Strip filesystem-unfriendly characters; we keep spaces because the
    # browser download dialog lets users rename anyway, and the source
    # name (e.g. "2026 Budget") reads nicer with a space.
    safe = re.sub(r"[\\/:*?\"<>|]", "_", plan_name)
    stamp = re.sub(r"[^0-9TZ:.-]", "", timestamp.replace("+00:00", ""))[:16]
    stamp = stamp.replace(":", "")
    return f"{safe} — exported {stamp}.xlsx"


def _inject_cached_values(
    rendered_bytes: bytes,
    sheet_name: str,
    values: dict[str, int | float],
) -> bytes:
    """Patch cached numeric values onto the formula cells in the rendered
    workbook so the importer's `_required_total_value` validation works.

    openpyxl writes formulas without any cached result, but `BudgetImporter`
    requires `cached_value` to be set on every "Total" row formula cell.
    This rewrites the underlying sheet XML in-place to inject `<v>` values.
    Same trick the test workbook builder uses (`tests/workbook_builder.py:212`).
    """
    if not values:
        return rendered_bytes
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir, "patched.xlsx")
        source_path = Path(temp_dir, "source.xlsx")
        source_path.write_bytes(rendered_bytes)
        with zipfile.ZipFile(source_path) as source_zip:
            source_zip.extractall(temp_dir)

        workbook_xml = ET.parse(Path(temp_dir, "xl", "workbook.xml"))
        rels_xml = ET.parse(Path(temp_dir, "xl", "_rels", "workbook.xml.rels"))
        relationship_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_xml.getroot().findall(f"{{{_NS_REL}}}Relationship")
        }
        target_path: str | None = None
        sheets_node = workbook_xml.getroot().find(f"{{{_NS_MAIN}}}sheets")
        if sheets_node is None:
            raise RuntimeError("Rendered workbook is missing <sheets> node.")
        for sheet in sheets_node:
            if sheet.attrib["name"] == sheet_name:
                target_path = relationship_map[sheet.attrib[f"{{{_NS_DOC_REL}}}id"]]
                break
        if target_path is None:
            raise RuntimeError(
                f"Sheet '{sheet_name}' not found in rendered workbook — "
                "exporter is misconfigured."
            )
        normalized = target_path.lstrip("/")
        if not normalized.startswith("xl/"):
            normalized = f"xl/{normalized}"
        sheet_path = Path(temp_dir, normalized)
        sheet_tree = ET.parse(sheet_path)
        cells = {
            cell.attrib["r"]: cell
            for cell in sheet_tree.getroot().findall(f".//{{{_NS_MAIN}}}c")
        }
        for ref, cached in values.items():
            cell = cells.get(ref)
            if cell is None:
                # Cell exists in the workbook (we wrote a formula there)
                # but the XML serializer may have written it under a
                # different row container — the dict lookup should be
                # exhaustive. Fail loud so a regression is obvious.
                raise RuntimeError(
                    f"Could not find cell {ref} in rendered sheet XML."
                )
            value_node = cell.find(f"{{{_NS_MAIN}}}v")
            if value_node is None:
                value_node = ET.SubElement(cell, f"{{{_NS_MAIN}}}v")
            value_node.text = str(cached)
        sheet_tree.write(sheet_path, encoding="utf-8", xml_declaration=True)

        with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as output_zip:
            for path in Path(temp_dir).rglob("*"):
                if path == temp_path or path == source_path or path.is_dir():
                    continue
                output_zip.write(path, path.relative_to(temp_dir))
        return temp_path.read_bytes()


def load_exported_workbook_for_test(content: bytes):
    """Test helper: load rendered xlsx bytes via openpyxl with formulas
    visible. Tests use this to assert cell values + formulas without
    re-running the importer."""
    return load_workbook(BytesIO(content), data_only=False)

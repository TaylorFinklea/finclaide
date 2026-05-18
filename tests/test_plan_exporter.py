from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from finclaide.budget_sheet import BudgetImporter
from finclaide.database import Database
from finclaide.plan_exporter import PlanExporter
from finclaide.plan_service import PlanService
from tests.workbook_builder import build_budget_workbook


def _setup_with_imported_plan(tmp_path: Path) -> tuple[Database, PlanExporter, BudgetImporter]:
    workbook_path = build_budget_workbook(tmp_path / "Budget.xlsx")
    database = Database(tmp_path / "finclaide.db")
    database.initialize()
    importer = BudgetImporter(database)
    importer.import_budget(workbook_path, "2026 Budget")
    plan_service = PlanService(database=database)
    exporter = PlanExporter(plan_service=plan_service)
    return database, exporter, importer


def test_exporter_produces_xlsx_with_block_columns_at_correct_offsets(tmp_path: Path):
    _, exporter, _ = _setup_with_imported_plan(tmp_path)
    result = exporter.export_active_plan(sheet_name="2026 Budget")
    workbook = load_workbook(BytesIO(result.bytes), data_only=True)
    sheet = workbook["2026 Budget"]
    # Monthly column A holds group + category names; B holds amounts.
    assert sheet["A2"].value == "Bills"  # group header
    assert sheet["B2"].value in (None, "")
    assert sheet["A3"].value == "Rent"
    assert float(sheet["B3"].value) == 1000.00
    # Yearly block column D + G monthly amounts.
    yearly_d_values = [sheet[f"D{r}"].value for r in range(2, 20)]
    assert "Yearly" in yearly_d_values
    # Stipends block at column I; Savings at column L.
    stipends_i = [sheet[f"I{r}"].value for r in range(1, 20)]
    savings_l = [sheet[f"L{r}"].value for r in range(1, 20)]
    assert "Stipends" in stipends_i
    assert "Savings" in savings_l


def test_exporter_round_trips_through_importer(tmp_path: Path):
    """Export, write to file, re-import — plan content (totals + counts)
    must round-trip identically. This is the load-bearing test for the
    importer-layout match."""
    database, exporter, importer = _setup_with_imported_plan(tmp_path)
    plan_service = PlanService(database=database)
    before = plan_service.get_active_plan()

    result = exporter.export_active_plan(sheet_name="2026 Budget")
    roundtrip_path = tmp_path / "Budget-roundtrip.xlsx"
    roundtrip_path.write_bytes(result.bytes)

    importer.import_budget(roundtrip_path, "2026 Budget")
    after = plan_service.get_active_plan()

    assert (
        before["totals"]["grand_total_milliunits"]
        == after["totals"]["grand_total_milliunits"]
    )
    for block in ("monthly", "annual", "one_time", "stipends", "savings"):
        assert len(before["blocks"][block]) == len(after["blocks"][block]), block
        before_pairs = sorted(
            (row["group_name"], row["category_name"], row["planned_milliunits"])
            for row in before["blocks"][block]
        )
        after_pairs = sorted(
            (row["group_name"], row["category_name"], row["planned_milliunits"])
            for row in after["blocks"][block]
        )
        assert before_pairs == after_pairs, block


def test_exporter_skips_ynab_system_categories(tmp_path: Path):
    database, exporter, _ = _setup_with_imported_plan(tmp_path)
    plan_service = PlanService(database=database)
    plan = plan_service.get_active_plan()
    before_row_count = exporter.export_active_plan(sheet_name="2026 Budget").row_count
    plan_service.create_category(
        plan["plan"]["id"],
        {
            "group_name": "Internal Master Category",
            "category_name": "Inflow: Ready to Assign",
            "block": "monthly",
            "planned_milliunits": 123000,
        },
    )

    result = exporter.export_active_plan(sheet_name="2026 Budget")
    workbook = load_workbook(BytesIO(result.bytes), data_only=True)
    sheet = workbook["2026 Budget"]
    values = [
        sheet[f"{column}{row}"].value
        for column in ("A", "D", "I", "L")
        for row in range(1, 53)
    ]

    assert result.row_count == before_row_count
    assert "Internal Master Category" not in values
    assert "Inflow: Ready to Assign" not in values


def test_exporter_emits_group_headers_with_empty_amount_cells(tmp_path: Path):
    _, exporter, _ = _setup_with_imported_plan(tmp_path)
    result = exporter.export_active_plan(sheet_name="2026 Budget")
    workbook = load_workbook(BytesIO(result.bytes), data_only=True)
    sheet = workbook["2026 Budget"]
    # Walk down column A and count empty B-cells where A has text — these
    # are group headers per the importer's contract.
    group_header_count = 0
    for row in range(2, 53):
        a = sheet[f"A{row}"].value
        b = sheet[f"B{row}"].value
        if a is not None and (b is None or b == ""):
            group_header_count += 1
    # The fixture imports rows from groups Bills + Expenses → at least 2 monthly groups.
    assert group_header_count >= 2


def test_exporter_writes_totals_row_with_formula(tmp_path: Path):
    _, exporter, _ = _setup_with_imported_plan(tmp_path)
    result = exporter.export_active_plan(sheet_name="2026 Budget")
    # data_only=False to see the formulas.
    workbook = load_workbook(BytesIO(result.bytes), data_only=False)
    sheet = workbook["2026 Budget"]
    assert sheet["A53"].value == "Total"
    assert isinstance(sheet["B53"].value, str) and sheet["B53"].value.startswith("=SUM(")
    assert sheet["G53"].value.startswith("=SUM(")


def test_exporter_filename_includes_plan_name_and_timestamp(tmp_path: Path):
    _, exporter, _ = _setup_with_imported_plan(tmp_path)
    result = exporter.export_active_plan(
        sheet_name="2026 Budget", now="2026-04-30T15:32:00+00:00"
    )
    assert result.filename.startswith("2026 Budget — exported ")
    assert result.filename.endswith(".xlsx")
    assert "2026-04-30" in result.filename


def test_exporter_handles_empty_savings_block(tmp_path: Path):
    """Removing all savings categories must not crash the exporter; the
    L:M columns are simply empty."""
    database, exporter, _ = _setup_with_imported_plan(tmp_path)
    with database.connect() as connection:
        connection.execute("DELETE FROM plan_categories WHERE block = 'savings'")
    result = exporter.export_active_plan(sheet_name="2026 Budget")
    workbook = load_workbook(BytesIO(result.bytes), data_only=True)
    sheet = workbook["2026 Budget"]
    # L1 still has the header; L2..L52 are empty.
    assert sheet["L1"].value == "Savings"
    assert sheet["L2"].value is None
